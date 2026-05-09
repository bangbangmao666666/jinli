"""Excel / CSV 解析器。

设计原则同 pdf_parser：只提取，不做语义判断。
支持 .xlsx / .xls（通过 openpyxl 或 xlrd）/ .csv。
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from ..errors import FileReadError, ParseError


def extract(file_path: str | Path) -> dict[str, Any]:
    """从 Excel/CSV 提取所有 Sheet 的原始行列数据。

    返回格式：
    {
        "file": "<path>",
        "file_type": "xlsx" | "xls" | "csv",
        "sheets": [
            {
                "name": "<sheet名>",
                "rows": [["v1", "v2", ...], ...],
                "row_count": N,
                "col_count": N,
                "header_row_index": <启发式猜测的表头行，可能为 null>
            }
        ]
    }
    """
    path = Path(file_path)
    if not path.exists():
        raise FileReadError(
            f"文件不存在：{path}",
            detail={"path": str(path)},
        )

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _extract_csv(path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _extract_xlsx(path)
    if suffix == ".xls":
        return _extract_xls(path)
    raise FileReadError(
        f"不支持的 Excel 格式：{suffix}。支持 .xlsx / .xls / .csv",
        detail={"path": str(path), "suffix": suffix},
    )


def _extract_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as e:
        raise FileReadError(
            "缺少依赖 openpyxl，请先运行 pip install openpyxl",
            detail={"hint": "pip install -r requirements.txt"},
        ) from e

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        raise FileReadError(
            f"无法打开 Excel 文件（可能已加密或损坏）：{exc}",
            detail={"path": str(path)},
        ) from exc

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [_clean_cell(c) for c in row]
            # 跳过全空行
            if any(v is not None for v in cleaned):
                rows.append(cleaned)

        if not rows:
            continue

        # 去掉尾部全空行
        while rows and all(v is None for v in rows[-1]):
            rows.pop()

        col_count = max(len(r) for r in rows) if rows else 0
        # 统一每行长度（补 None）
        rows = [r + [None] * (col_count - len(r)) for r in rows]

        sheets.append({
            "name": sheet_name,
            "rows": rows,
            "row_count": len(rows),
            "col_count": col_count,
            "header_row_index": _guess_header_row(rows),
        })

    wb.close()

    if not sheets:
        raise ParseError(
            "Excel 文件中没有找到任何有效数据（所有 Sheet 均为空）",
            detail={"path": str(path)},
        )

    return {
        "file": str(path),
        "file_type": "xlsx",
        "sheets": sheets,
    }


def _extract_xls(path: Path) -> dict[str, Any]:
    try:
        import xlrd
    except ImportError as e:
        raise FileReadError(
            "读取 .xls 格式需要 xlrd，请运行 pip install xlrd>=2.0",
            detail={"hint": "或将文件另存为 .xlsx 格式后重新上传"},
        ) from e

    try:
        wb = xlrd.open_workbook(str(path))
    except Exception as exc:
        raise FileReadError(f"无法打开 xls 文件：{exc}", detail={"path": str(path)}) from exc

    sheets = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for rx in range(ws.nrows):
            cleaned = [_clean_cell(ws.cell_value(rx, cx)) for cx in range(ws.ncols)]
            if any(v is not None for v in cleaned):
                rows.append(cleaned)
        if rows:
            sheets.append({
                "name": sheet_name,
                "rows": rows,
                "row_count": len(rows),
                "col_count": ws.ncols,
                "header_row_index": _guess_header_row(rows),
            })

    return {"file": str(path), "file_type": "xls", "sheets": sheets}


def _extract_csv(path: Path) -> dict[str, Any]:
    # 尝试常见编码
    for enc in ("utf-8-sig", "gbk", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                rows = []
                for row in reader:
                    cleaned = [_clean_cell(c) for c in row]
                    if any(v is not None for v in cleaned):
                        rows.append(cleaned)
            if rows:
                return {
                    "file": str(path),
                    "file_type": "csv",
                    "sheets": [{
                        "name": "Sheet1",
                        "rows": rows,
                        "row_count": len(rows),
                        "col_count": max(len(r) for r in rows),
                        "header_row_index": _guess_header_row(rows),
                    }],
                }
        except (UnicodeDecodeError, csv.Error):
            continue

    raise ParseError(
        f"无法读取 CSV 文件（尝试了 utf-8/gbk/gb18030 编码均失败）：{path.name}",
        detail={"path": str(path), "hint": "请确认文件编码，或转存为 xlsx 后重新上传"},
    )


# ─── 辅助函数 ────────────────────────────────────────────────────────────────

# 表头行常见关键词（用于启发式猜测 header_row_index）
_HEADER_KEYWORDS = [
    "院校代码", "学校代码", "院校名称", "专业代码", "专业名称",
    "选考科目", "选科要求", "计划人数", "招生人数",
    "最低分", "最低位次", "省控线", "年份", "批次",
]


def _guess_header_row(rows: list) -> int | None:
    """启发式：找第一行包含表头关键词最多的行。"""
    best_row, best_count = 0, 0
    for i, row in enumerate(rows[:10]):  # 只看前10行
        count = sum(
            1 for cell in row
            if cell and any(kw in str(cell) for kw in _HEADER_KEYWORDS)
        )
        if count > best_count:
            best_count, best_row = count, i
    return best_row if best_count > 0 else None


def _clean_cell(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    # 处理 openpyxl 读出来的浮点院校代码（如 10001.0 → "10001"）
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    import re
    s = re.sub(r"\s+", " ", s)
    return s if s else None
