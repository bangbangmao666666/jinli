"""数据准备 单元测试。

由于真实 PDF 不在代码仓库中，PDF 相关测试使用 mock；
Excel 测试用 openpyxl 动态生成临时文件（不依赖外部文件）。
"""
import json
import sys
import tempfile
import unittest
from pathlib import Path

# sys.path 设置
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from data_模块.normalizer import normalize_admission_plan, _detect_column_map
from data_模块.validator import validate
from data_模块.errors import NormalizationError, FileReadError, ParseError


# ─── normalizer 测试 ──────────────────────────────────────────────────────────

SAMPLE_TABLES_FLAT = [
    {
        "page": 1,
        "table_index": 0,
        "rows": [
            # 表头行
            ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求", "计划人数"],
            # 数据行
            ["10459", "郑州大学", "01", "080601", "电气工程及其自动化", "物理+化学", "60"],
            ["10459", "郑州大学", "01", "081301", "化学工程与工艺",     "物理+化学", "30"],
            ["10459", "郑州大学", "02", "050101", "汉语言文学",         "不限",      "50"],
            ["10698", "西安交通大学", "01", "080901", "计算机科学与技术", "物理+化学", "40"],
        ],
    }
]

SAMPLE_RAW = {
    "file": "/tmp/河南2026招生计划.pdf",
    "total_pages": 5,
    "is_likely_scanned": False,
    "pages": [{"page": 1, "text_snippet": "2026年河南省普通高校招生计划", "tables": []}],
    "all_tables_flat": SAMPLE_TABLES_FLAT,
}


class TestColumnDetect(unittest.TestCase):
    def test_standard_header_detected(self):
        header = ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求"]
        cm = _detect_column_map(header)
        self.assertEqual(cm["school_code"], 0)
        self.assertEqual(cm["school_name"], 1)
        self.assertEqual(cm["group_code"], 2)
        self.assertEqual(cm["major_code"], 3)
        self.assertEqual(cm["major_name"], 4)
        self.assertEqual(cm["subject_req_raw"], 5)

    def test_alternate_header_keywords(self):
        header = ["学校代码", "学校名称", "组号", "专业号", "专业"]
        cm = _detect_column_map(header)
        self.assertIn("school_code", cm)
        self.assertIn("school_name", cm)
        self.assertIn("major_name", cm)


class TestNormalizer(unittest.TestCase):
    def test_basic_normalization(self):
        result = normalize_admission_plan(SAMPLE_RAW)
        self.assertEqual(result["province"], "河南")
        self.assertEqual(result["valid_for_year"], 2026)
        self.assertEqual(len(result["schools"]), 2)

    def test_school_structure(self):
        result = normalize_admission_plan(SAMPLE_RAW)
        zzu = next(s for s in result["schools"] if s["school_code"] == "10459")
        self.assertEqual(zzu["school_name"], "郑州大学")
        self.assertEqual(len(zzu["major_groups"]), 2)

    def test_subject_requirement_raw_untouched(self):
        """选科要求原文绝对不能被修改。"""
        result = normalize_admission_plan(SAMPLE_RAW)
        zzu = next(s for s in result["schools"] if s["school_code"] == "10459")
        group_01 = next(g for g in zzu["major_groups"] if g["group_code"] == "01")
        # 必须与原文完全一致，包括空格和符号
        self.assertEqual(group_01["subject_requirement_raw"], "物理+化学")

    def test_source_type_is_user_upload(self):
        result = normalize_admission_plan(SAMPLE_RAW)
        self.assertEqual(result["source_type"], "user_upload")

    def test_user_hint_overrides_province(self):
        result = normalize_admission_plan(SAMPLE_RAW, user_hint={"province": "江苏"})
        self.assertEqual(result["province"], "江苏")

    def test_empty_tables_raises(self):
        raw_empty = {**SAMPLE_RAW, "all_tables_flat": []}
        with self.assertRaises(NormalizationError):
            normalize_admission_plan(raw_empty)

    def test_yearly_admission_data_empty_by_default(self):
        """招生计划文件不包含历史位次，该字段应该是空列表。"""
        result = normalize_admission_plan(SAMPLE_RAW)
        for school in result["schools"]:
            for group in school["major_groups"]:
                for major in group["majors"]:
                    self.assertEqual(major["yearly_admission_data"], [])

    def test_military_school_detected(self):
        raw = {**SAMPLE_RAW, "all_tables_flat": [{
            "page": 1, "table_index": 0,
            "rows": [
                ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求"],
                ["90001", "解放军信息工程大学", "01", "080601", "电气工程", "物理+化学"],
            ]
        }]}
        result = normalize_admission_plan(raw)
        school = result["schools"][0]
        self.assertTrue(school["is_military"])

    def test_police_school_detected(self):
        raw = {**SAMPLE_RAW, "all_tables_flat": [{
            "page": 1, "table_index": 0,
            "rows": [
                ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求"],
                ["10041", "中国人民公安大学", "01", "030602", "治安学", "物理或化学"],
            ]
        }]}
        result = normalize_admission_plan(raw)
        school = result["schools"][0]
        self.assertTrue(school["is_police"])


class TestValidator(unittest.TestCase):
    def _make_valid(self):
        return normalize_admission_plan(SAMPLE_RAW)

    def test_valid_plan_passes(self):
        normalized = self._make_valid()
        report = validate(normalized)
        self.assertTrue(report["passed"])
        self.assertEqual(len(report["errors"]), 0)

    def test_missing_province_is_error(self):
        normalized = self._make_valid()
        normalized["province"] = None
        report = validate(normalized)
        self.assertFalse(report["passed"])
        self.assertTrue(any("province" in e for e in report["errors"]))

    def test_missing_year_is_error(self):
        normalized = self._make_valid()
        normalized["valid_for_year"] = None
        report = validate(normalized)
        self.assertFalse(report["passed"])
        self.assertTrue(any("valid_for_year" in e for e in report["errors"]))

    def test_user_upload_warning_always_present(self):
        normalized = self._make_valid()
        report = validate(normalized)
        user_upload_warnings = [w for w in report["warnings"] if "用户上传" in w]
        self.assertGreater(len(user_upload_warnings), 0)

    def test_stats_correct(self):
        normalized = self._make_valid()
        report = validate(normalized)
        self.assertEqual(report["stats"]["total_schools"], 2)


# ─── Excel parser 集成测试（动态生成临时文件）───────────────────────────────

class TestExcelParser(unittest.TestCase):
    def _make_xlsx(self, rows: list) -> Path:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "招生计划"
        for row in rows:
            ws.append(row)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        return Path(tmp.name)

    def test_extract_basic_xlsx(self):
        from data_模块.parsers.excel_parser import extract
        path = self._make_xlsx([
            ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求"],
            ["10459", "郑州大学", "01", "080601", "电气工程及其自动化", "物理+化学"],
        ])
        result = extract(str(path))
        self.assertEqual(result["file_type"], "xlsx")
        self.assertEqual(len(result["sheets"]), 1)
        sheet = result["sheets"][0]
        self.assertEqual(sheet["row_count"], 2)

    def test_excel_to_normalized(self):
        from data_模块.parsers.excel_parser import extract
        path = self._make_xlsx([
            ["院校代码", "院校名称", "专业组代码", "专业代码", "专业名称", "选考科目要求"],
            ["10459", "郑州大学2026", "01", "080601", "电气工程及其自动化", "物理+化学"],
        ])
        raw = extract(str(path))
        result = normalize_admission_plan(
            raw, user_hint={"province": "河南", "valid_for_year": 2026}
        )
        self.assertEqual(result["province"], "河南")
        self.assertEqual(len(result["schools"]), 1)

    def test_float_school_code_normalized(self):
        """openpyxl 可能把 10459 读成 10459.0，要确保被规范化为字符串。"""
        from data_模块.parsers.excel_parser import extract
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "院校代码"
        ws["B1"] = "院校名称"
        ws["C1"] = "专业组代码"
        ws["D1"] = "专业代码"
        ws["E1"] = "专业名称"
        ws["F1"] = "选考科目要求"
        ws["A2"] = 10459   # 整数，会被 openpyxl 读出再转字符串
        ws["B2"] = "郑州大学"
        ws["C2"] = "01"
        ws["D2"] = 80601
        ws["E2"] = "电气工程及其自动化"
        ws["F2"] = "物理+化学"
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        raw = extract(tmp.name)
        result = normalize_admission_plan(
            raw, user_hint={"province": "河南", "valid_for_year": 2026}
        )
        sc = result["schools"][0]["school_code"]
        self.assertNotIn(".0", sc)


class TestPdfParserMocked(unittest.TestCase):
    """PDF parser 不依赖真实文件，只测试错误处理路径。"""

    def test_nonexistent_file_raises(self):
        from data_模块.parsers.pdf_parser import extract
        with self.assertRaises(FileReadError):
            extract("/nonexistent/path/file.pdf")

    def test_wrong_extension_raises(self):
        from data_模块.parsers.pdf_parser import extract
        with self.assertRaises(FileReadError):
            extract("/tmp/test.xlsx")


if __name__ == "__main__":
    unittest.main()
